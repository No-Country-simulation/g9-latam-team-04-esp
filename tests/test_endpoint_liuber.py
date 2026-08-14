"""
Corre los 43 registros de la pestana "Liuber" de Regitros_TEST_ENDPOINT_ENG_ESP.xlsx
contra el modelo de clasificacion (EN y ES), y llena "Registro Usado" y
"Observacion_Resutado_Test" con el resultado.

No toca las demas pestanas (Jhonny, Gabriel, Henry, Roman, Luca, Alan, Jaime).
Guarda el resultado en una copia nueva, no pisa el archivo original.
"""

import sys
from pathlib import Path

REPO_ROOT = Path(r"C:\Users\sinof\g9-latam-team-04-esp")
sys.path.insert(0, str(REPO_ROOT))

import openpyxl

from backend.src.services.clasificador import clasificador

ARCHIVO_ORIGEN = Path(r"C:\Users\sinof\Downloads\Regitros_TEST_ENDPOINT_ENG_ESP.xlsx")
ARCHIVO_SALIDA = Path(r"C:\Users\sinof\Downloads\Regitros_TEST_ENDPOINT_ENG_ESP_Liuber_resultados.xlsx")

HOJA = "Liuber"
FILA_INICIO = 4
FILA_FIN = 46

# Columnas (1-indexed) segun el header de la fila 3
COL_ITEM = 1
COL_REGISTRO_USADO = 2
COL_OBSERVACION = 3
COL_TITULO = 4
COL_TEXTO = 5
COL_CATEGORIA = 6
COL_TITULO_ESP = 8
COL_TEXTO_ESP = 9


def evaluar(titulo: str, texto: str, idioma: str, categoria_esperada: str) -> tuple[bool, str]:
    resultado = clasificador.predecir(titulo, texto, idioma=idioma)
    predicha = resultado["categoria"].strip().lower()
    esperada = categoria_esperada.strip().lower()
    confianza = resultado["probabilidad"]

    ok = predicha == esperada
    if ok:
        texto_resultado = f"OK ({predicha}, {confianza:.1%})"
    else:
        texto_resultado = f"FALLO - predijo '{predicha}' ({confianza:.1%}), esperado '{esperada}'"
    return ok, texto_resultado


def main():
    print("Cargando modelo...")
    clasificador.cargar()
    if not clasificador.cargado:
        print("ERROR: no se pudo cargar el modelo. Revisa data-science/models/en|es/.")
        return

    print(f"Abriendo {ARCHIVO_ORIGEN.name}...")
    wb = openpyxl.load_workbook(ARCHIVO_ORIGEN)
    ws = wb[HOJA]

    total = 0
    ok_en = ok_es = 0

    for fila in range(FILA_INICIO, FILA_FIN + 1):
        titulo = ws.cell(row=fila, column=COL_TITULO).value
        texto = ws.cell(row=fila, column=COL_TEXTO).value
        categoria = ws.cell(row=fila, column=COL_CATEGORIA).value
        titulo_esp = ws.cell(row=fila, column=COL_TITULO_ESP).value
        texto_esp = ws.cell(row=fila, column=COL_TEXTO_ESP).value

        if not titulo or not texto or not categoria:
            continue

        total += 1

        ok1, obs_en = evaluar(titulo, texto, "en", categoria)
        ok_en += ok1

        if titulo_esp and texto_esp:
            ok2, obs_es = evaluar(titulo_esp, texto_esp, "es", categoria)
            ok_es += ok2
            observacion = f"EN: {obs_en} | ES: {obs_es}"
        else:
            observacion = f"EN: {obs_en} | ES: sin texto traducido"

        ws.cell(row=fila, column=COL_REGISTRO_USADO, value="Si")
        ws.cell(row=fila, column=COL_OBSERVACION, value=observacion)

        print(f"  Item {ws.cell(row=fila, column=COL_ITEM).value}: {observacion}")

    wb.save(ARCHIVO_SALIDA)

    print()
    print(f"Total evaluados: {total}")
    print(f"Aciertos EN: {ok_en}/{total} ({ok_en/total:.1%})")
    print(f"Aciertos ES: {ok_es}/{total} ({ok_es/total:.1%})")
    print(f"Guardado en: {ARCHIVO_SALIDA}")


if __name__ == "__main__":
    main()
